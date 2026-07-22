#!/usr/bin/env python3
"""Generate Table A – Go/No-Go Metrics for the EM AI proposal."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Table A – Go-NoGo Metrics"

# ── Column widths ──
col_widths = [38, 38, 42, 30, 42]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
cell_font = Font(name="Calibri", size=11)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Headers ──
headers = [
    "Metric",
    "What is Measured",
    "Measurement Approach",
    "Go/No-Go Threshold",
    "Relevance to AI Advantage",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

# ── Data rows (derived from proposal milestones & metrics) ──
rows = [
    # 1 – Evidence Retrieval Performance (with baseline comparison)
    [
        "Evidence Retrieval Precision (Precision@k)",
        "Fraction of top-k retrieved passages that are relevant to an EM decision query, measured across air, soil, and water domains, compared against keyword-search and manual-retrieval baselines",
        "Submit a benchmark set of ≥50 EM decision questions to the mRAG pipeline and to a keyword-search baseline (BM25); domain experts label top-10 retrieved passages as relevant/irrelevant; compute Precision@10 for both systems",
        "Precision@10 ≥ 0.80; ≥ 30% relative improvement over BM25 keyword-search baseline",
        "Demonstrates that AI-guided multimodal retrieval (mRAG with domain-adapted embeddings) surfaces verifiable, decision-relevant evidence that keyword search alone cannot discover, enabling cross-source synthesis previously infeasible at scale",
    ],
    # 2 – Cross-Domain Schema Alignment (with baseline comparison)
    [
        "Cross-Domain Schema Alignment Rate",
        "Percentage of variables in a new environmental medium (water or soil) that are automatically mapped to the common harmonization schema without manual correction, versus a keyword-matching NLP baseline",
        "Apply the air-domain harmonization schema to candidate water/soil datasets using the AI alignment engine and a rule-based keyword-matching baseline; count variables auto-aligned vs. total variables requiring alignment for each approach",
        "≥ 85% auto-alignment; ≥ 3× improvement over keyword-matching baseline (expected < 30% auto-alignment)",
        "Validates that AI-guided semantic schema alignment generalizes across EM media, reducing months of manual harmonization to hours and enabling rapid onboarding of new environmental domains",
    ],
    # 3 – Predictive Analytics Accuracy (with explicit baseline improvement)
    [
        "Air-Domain Predictive Accuracy (RMSE / R²)",
        "Forecast accuracy for selected air-quality indicators (e.g., temperature, wind speed, contamination proxy) using ML models trained on the harmonized repository, compared against persistence and climatology baselines",
        "Train ML models on 80% of historical ORR meteorological records; evaluate on held-out 20%; compute RMSE and R² for each target variable; compare against persistence (last-value) and climatological-mean baselines",
        "R² ≥ 0.75 for ≥2 target variables; ≥ 20% RMSE reduction over persistence/climatology baseline",
        "Confirms that AI-enhanced analytics produce actionable early-warning forecasts that significantly outperform conventional statistical approaches, transforming EM monitoring from reactive to predictive",
    ],
    # 4 – Data Quality & Completeness (operationalized sub-scores)
    [
        "Repository Data Quality Pass Rate",
        "Percentage of ingested records that pass all six independently defined QA checks: (1) completeness ≥ 90%, (2) unit consistency, (3) spatial coordinate validity, (4) temporal coherence (no future dates, monotonic sequences), (5) provenance traceability (linked to source document), (6) deduplication (no exact or near-duplicates)",
        "Run the automated quality engine on every ingested record; each record is independently evaluated against the six binary QA checks; compute the fraction of records passing all six checks",
        "≥ 95% of air-domain records pass all 6 QA checks; ≥ 85% of initial soil/water records pass all 6 checks",
        "AI-automated quality validation replaces weeks of manual QA review, ensures that downstream ML models and decision-support outputs are grounded in verified data, and provides reproducible, auditable quality evidence for regulatory compliance",
    ],
    # 5 – Anomaly / Early-Warning Detection (unchanged — already strong)
    [
        "Anomaly Detection Recall (Early Warning)",
        "Proportion of known historical contamination events or extreme-weather episodes correctly flagged by the early-warning analytics module, validated against pre-1999 expert-documented ground truth",
        "Compile a ground-truth set of ≥20 documented anomalous events from pre-1999 ORR records; run the AI detection module and a threshold-based statistical baseline; compute recall and false-positive rate for both",
        "Recall ≥ 0.80 with false-positive rate ≤ 0.20; ≥ 25% recall improvement over threshold-based statistical baseline",
        "AI-driven anomaly detection identifies complex, multivariate contamination signatures that fixed-threshold methods miss, transforming EM early-warning from univariate alarms to context-aware, multi-parameter surveillance",
    ],
    # 6 – HIL Pipeline End-to-End Fidelity (operationalized provenance checklist)
    [
        "HIL Pipeline End-to-End Fidelity",
        "Whether the supervisor agent, knowledge integration, and conflict-resolution layers complete a full decision scenario with provenance-linked outputs (defined as: every output linked to ≥1 source document with timestamp, version, confidence score, and CoT trace) and a completed HIL escalation cycle",
        "Execute ≥3 representative EM decision scenarios (characterization, monitoring, remediation) end-to-end; verify each output against a 5-item provenance checklist (source link, timestamp, version, confidence, CoT trace); measure escalation agreement via Cohen's kappa between system and expert panel (≥2 experts)",
        "100% of outputs satisfy all 5 provenance checklist items; escalation Cohen's κ ≥ 0.80 between system and expert panel",
        "Demonstrates that multi-agent AI orchestration maintains full accountability and human oversight with measurable inter-rater reliability, critical for regulatory compliance and establishing trust in AI-assisted EM decision-making at DOE sites",
    ],
    # 7 – Workflow Acceleration (strengthened threshold)
    [
        "Overall Workflow Acceleration",
        "Wall-clock time from raw data ingestion to decision-ready analytic output using the AI pipeline, compared to the conventional manual expert workflow for the same task",
        "Time ≥5 representative retrieval-harmonization-analysis-visualization sequences performed by a domain expert (manual baseline) and by the AI pipeline; compute median speedup ratio across all sequences",
        "≥ 10× median acceleration over manual baseline",
        "Order-of-magnitude acceleration transforms EM decision timelines from weeks to hours, enabling near-real-time adaptive monitoring and rapid response to emerging contamination events — a capability impossible without AI-guided automation",
    ],
    # 8 – Decision Dashboard Deployment Time (reframed from sketch-to-code)
    [
        "Decision Dashboard Deployment Time",
        "Elapsed time from receipt of a new EM data domain or stakeholder visualization request to a deployed, functional, interactive decision dashboard, using the AI sketch-to-code pipeline",
        "For ≥5 representative new-domain requests (covering geospatial overlays, temporal trends, what-if panels, and cross-domain views), measure wall-clock time from SVG sketch to deployed dashboard; compare against historical manual dashboard development timelines",
        "≤ 4 hours from sketch to deployed dashboard (vs. ≥ 2 weeks manual baseline); ≥ 80% first-pass rendering success rate",
        "AI-driven interface generation reduces dashboard deployment from weeks to hours, enabling rapid adaptation to new EM sites and data domains and empowering non-technical stakeholders to interact with AI-synthesized insights directly",
    ],
    # 9 – Embedding Drift Detection Latency (NEW — knowledge graph innovation)
    [
        "Embedding Drift Detection Latency",
        "Time elapsed between the onset of a distributional shift in incoming data streams (relative to baseline embeddings) and the system's automated re-indexing trigger and agent recalibration flag",
        "Inject ≥10 synthetic distributional shifts of known magnitude (concept drift, covariate shift) into the data pipeline at recorded timestamps; measure the delay between injection and automated re-indexing trigger; verify that flagged outlier assessments are routed for human review",
        "Detection latency ≤ 24 hours from shift onset to re-indexing trigger for ≥ 90% of injected shifts; 100% of flagged outliers routed to human review queue",
        "Continuous, automated drift monitoring ensures the knowledge graph and retrieval layer remain calibrated to evolving site conditions — a self-correcting AI capability that prevents stale or misleading decision-support outputs without manual surveillance",
    ],
    # 10 – Feedback-Driven Output Improvement (NEW — DPO/RLHF loop)
    [
        "Feedback-Driven Output Preference Win-Rate (DPO/RLHF)",
        "Proportion of pairwise comparisons in which domain experts prefer post-optimization system outputs over pre-optimization outputs, measuring the effectiveness of supervised fine-tuning, DPO, and RLHF orchestration updates",
        "Collect ≥50 pre-optimization and post-optimization output pairs across representative EM decision tasks; present pairs (blinded, randomized order) to ≥3 domain experts; compute win-rate and 95% confidence interval; verify no catastrophic forgetting via regression tests on held-out benchmark set",
        "Post-optimization win-rate ≥ 70% (95% CI lower bound > 55%); zero regression failures on held-out benchmark set",
        "Demonstrates that the AI system measurably improves with human feedback over time — a transformative closed-loop capability where expert corrections are systematically converted into better decision-support outputs, compounding value with each deployment cycle",
    ],
]

for r, row_data in enumerate(rows, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = cell_font
        cell.alignment = wrap
        cell.border = thin_border

# ── Auto-fit row heights (approximate) ──
for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
    max_lines = 1
    for cell in row:
        if cell.value:
            col_w = col_widths[cell.column - 1]
            lines = len(str(cell.value)) / (col_w * 1.2)  # rough char-per-line
            max_lines = max(max_lines, lines)
    ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

out = "/home/jose/nsrd_ornl/nsrd_ui/docs/Table_A_GoNoGo_Metrics.xlsx"
wb.save(out)
print(f"Saved → {out}")
